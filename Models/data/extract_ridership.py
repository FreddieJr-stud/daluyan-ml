"""Extract ridership data from the Summarized Ridership Excel file.

Parses the Excel file, computes passengers on board per trip leg,
matches to segment IDs in DuckDB, and inserts into segment_ridership table.

Two data sources in the file:
  1. Summarized sheet (29 detailed days): per-station boarding/alighting
     with timestamps → exact passengers per leg.
  2. Total-only sheet (30 dates): daily direction totals only →
     distributed using average boarding pattern from detailed days.

Station name mapping:
  Ridership "Quinta" = system "Quinta" (canonical station name)

Usage:
    python -m data.extract_ridership
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import duckdb
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import DUCKDB_PATH, MAX_PASSENGERS

# ── Constants ─────────────────────────────────────────────────────────────

EXCEL_PATH = Path(__file__).resolve().parent.parent / "[Twinning] Summarized Ridership Data.xlsx"

# Map ridership station names → system station names
STATION_MAP = {
    "Guadalupe": "Guadalupe",
    "Hulo": "Hulo",
    "Lambingan": "Lambingan",
    "PUP": "PUP",
    "Quinta": "Quinta",    # Quinta Market station
    "Escolta": "Escolta",
}

# Downstream station order (system names)
DOWNSTREAM_ORDER = ["Guadalupe", "Hulo", "Lambingan", "PUP", "Quinta", "Escolta"]
# Upstream station order (system names)
UPSTREAM_ORDER = ["Escolta", "Quinta", "PUP", "Lambingan", "Hulo", "Guadalupe"]


# ── Data Structures ───────────────────────────────────────────────────────

class StopRecord:
    """One stop on a trip: station + boarding/alighting counts."""
    __slots__ = ("station", "boarding", "alighting")

    def __init__(self, station: str, boarding: float, alighting: float):
        self.station = station
        self.boarding = boarding
        self.alighting = alighting


class LegRecord:
    """Passengers on board between two stations."""
    __slots__ = ("departure", "arrival", "passengers_on_board")

    def __init__(self, departure: str, arrival: str, passengers_on_board: int):
        self.departure = departure
        self.arrival = arrival
        self.passengers_on_board = passengers_on_board


# ── Parsing ───────────────────────────────────────────────────────────────

def _fix_year(d: datetime) -> datetime:
    """Fix Nov/Dec dates erroneously stored as year 2026 → 2025."""
    if d.month in (11, 12) and d.year == 2026:
        return d.replace(year=2025)
    return d


def _parse_summarized_sheet(wb: openpyxl.Workbook) -> dict[str, list[StopRecord]]:
    """Parse the Summarized sheet into per-date, per-direction stop records.

    Returns: {
        "2025-11-19_downstream": [StopRecord(Guadalupe, 27, 0), ...],
        "2025-11-19_upstream": [StopRecord(Escolta, 5, 0), ...],
    }
    """
    ws = wb["Summarized"]
    trips: dict[str, list[StopRecord]] = {}

    # Group rows by date
    day_rows: dict[str, list[tuple]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if d is None or not hasattr(d, "strftime"):
            continue
        d = _fix_year(d)
        date_str = d.strftime("%Y-%m-%d")
        station = (row[2] or "").strip()
        if not station:
            continue

        # Map station name
        sys_station = STATION_MAP.get(station)
        if sys_station is None:
            continue  # Unknown station

        loading = float(row[5] or 0)
        alighting = float(row[6] or 0)

        if date_str not in day_rows:
            day_rows[date_str] = []
        day_rows[date_str].append((sys_station, loading, alighting))

    # Split each day into downstream + upstream using station order
    for date_str, stops in day_rows.items():
        n = len(stops)
        if n < 2:
            continue

        # First half = downstream (or first trip), second half = upstream (return trip)
        mid = n // 2
        downstream_stops = stops[:mid]
        upstream_stops = stops[mid:]

        # Validate by checking station sequences
        ds_key = f"{date_str}_downstream"
        us_key = f"{date_str}_upstream"

        trips[ds_key] = [StopRecord(s, b, a) for s, b, a in downstream_stops]
        trips[us_key] = [StopRecord(s, b, a) for s, b, a in upstream_stops]

    return trips


def _compute_legs(stops: list[StopRecord]) -> list[LegRecord]:
    """Convert stop records into leg records with cumulative passenger count."""
    legs = []
    pax = 0
    for i, stop in enumerate(stops):
        pax += stop.boarding - stop.alighting
        pax = max(pax, 0)  # Safety: can't have negative passengers

        # This leg is from current stop to next stop
        if i < len(stops) - 1:
            next_station = stops[i + 1].station
            legs.append(LegRecord(stop.station, next_station, int(round(pax))))

    return legs


def _compute_boarding_distribution(
    trips: dict[str, list[StopRecord]],
) -> dict[str, dict[str, float]]:
    """Compute average boarding/alighting fractions per station per direction.

    Returns: {
        "downstream": {"Guadalupe": (0.82, 0.0), "Hulo": (0.15, 0.02), ...},
        "upstream": {"Escolta": (0.48, 0.0), ...},
    }
    Each value is (boarding_fraction, alighting_fraction) of direction total.
    """
    totals: dict[str, dict[str, list[float]]] = {
        "downstream": {},
        "upstream": {},
    }

    for key, stops in trips.items():
        direction = key.split("_", 1)[1]  # "downstream" or "upstream"
        if direction not in totals:
            continue

        day_boarding = sum(s.boarding for s in stops)
        day_alighting = sum(s.alighting for s in stops)

        for stop in stops:
            if stop.station not in totals[direction]:
                totals[direction][stop.station] = []
            # Store (boarding, alighting, total_boarding, total_alighting)
            totals[direction][stop.station].append(
                (stop.boarding, stop.alighting, day_boarding, day_alighting)
            )

    # Compute averages
    distributions: dict[str, dict[str, tuple[float, float]]] = {}
    for direction, stations in totals.items():
        distributions[direction] = {}
        for station, records in stations.items():
            total_b = sum(r[2] for r in records)
            total_a = sum(r[3] for r in records)
            sum_b = sum(r[0] for r in records)
            sum_a = sum(r[1] for r in records)

            b_frac = sum_b / total_b if total_b > 0 else 0
            a_frac = sum_a / total_a if total_a > 0 else 0
            distributions[direction][station] = (b_frac, a_frac)

    return distributions


def _parse_total_only_sheet(
    wb: openpyxl.Workbook,
    distributions: dict[str, dict[str, tuple[float, float]]],
) -> dict[str, list[StopRecord]]:
    """Parse Total-only sheet and distribute using average boarding pattern.

    Returns same format as _parse_summarized_sheet.
    """
    ws = wb["Total only (Jan 14 - Feb 26)"]
    trips: dict[str, list[StopRecord]] = {}

    for row in ws.iter_rows(values_only=True):
        d = row[0]
        if d is None or not hasattr(d, "strftime"):
            continue
        date_str = d.strftime("%Y-%m-%d")
        ds_total = float(row[1] or 0)  # downstream boarding total
        us_total = float(row[2] or 0)  # upstream boarding total

        # Distribute downstream
        ds_stops = []
        for station in DOWNSTREAM_ORDER:
            b_frac, a_frac = distributions.get("downstream", {}).get(station, (0, 0))
            boarding = round(ds_total * b_frac)
            alighting = round(ds_total * a_frac)
            ds_stops.append(StopRecord(station, boarding, alighting))
        trips[f"{date_str}_downstream"] = ds_stops

        # Distribute upstream
        us_stops = []
        for station in UPSTREAM_ORDER:
            b_frac, a_frac = distributions.get("upstream", {}).get(station, (0, 0))
            boarding = round(us_total * b_frac)
            alighting = round(us_total * a_frac)
            us_stops.append(StopRecord(station, boarding, alighting))
        trips[f"{date_str}_upstream"] = us_stops

    return trips


# ── DuckDB Integration ────────────────────────────────────────────────────

def _get_segments_by_date_direction(db: duckdb.DuckDBPyConnection) -> dict[str, list[dict]]:
    """Get all segments grouped by date + direction.

    Returns: {"2025-11-19_downstream": [{"segment_id": ..., "departure_station": ...}, ...]}
    """
    rows = db.execute("""
        SELECT segment_id, departure_station, arrival_station, direction,
               CAST(departure_time AS DATE) as dep_date, departure_time
        FROM segments
        WHERE direction != 'unknown'
          AND departure_station != 'unknown'
        ORDER BY departure_time
    """).fetchall()

    groups: dict[str, list[dict]] = {}
    for seg_id, dep, arr, direction, dep_date, dep_time in rows:
        key = f"{dep_date}_{direction}"
        if key not in groups:
            groups[key] = []
        groups[key].append({
            "segment_id": seg_id,
            "departure_station": dep,
            "arrival_station": arr,
        })

    return groups


def _match_and_insert(
    db: duckdb.DuckDBPyConnection,
    all_trips: dict[str, list[StopRecord]],
    segments_by_key: dict[str, list[dict]],
    source_label: str,
    verbose: bool = True,
) -> tuple[int, int]:
    """Match ridership legs to segments and insert into segment_ridership.

    Returns (matched, unmatched) counts.
    """
    matched = 0
    unmatched = 0

    for trip_key, stops in all_trips.items():
        legs = _compute_legs(stops)
        if not legs:
            continue

        # Build lookup: departure_station → passengers_on_board
        pax_by_dep = {leg.departure: leg.passengers_on_board for leg in legs}

        # Find matching segments
        segments = segments_by_key.get(trip_key, [])
        for seg in segments:
            dep_station = seg["departure_station"]
            if dep_station in pax_by_dep:
                pob = pax_by_dep[dep_station]
                load_ratio = pob / MAX_PASSENGERS

                db.execute("""
                    INSERT OR REPLACE INTO segment_ridership
                        (segment_id, passengers_on_board, passenger_load_ratio)
                    VALUES (?, ?, ?)
                """, [seg["segment_id"], pob, load_ratio])
                matched += 1
            else:
                unmatched += 1

    if verbose:
        print(f"  {source_label}: {matched} matched, {unmatched} unmatched")

    return matched, unmatched


# ── Main ──────────────────────────────────────────────────────────────────

def extract(verbose: bool = True) -> dict:
    """Parse ridership Excel and populate segment_ridership in DuckDB."""
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    if verbose:
        print(f"Reading ridership data from: {EXCEL_PATH.name}")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Step 1: Parse detailed days (29 dates, per-station data)
    detailed_trips = _parse_summarized_sheet(wb)
    n_detailed_dates = len(set(k.rsplit("_", 1)[0] for k in detailed_trips))
    if verbose:
        print(f"\nDetailed days: {n_detailed_dates} dates, "
              f"{len(detailed_trips)} trips (downstream + upstream)")

    # Step 2: Compute average boarding distribution for distributing totals
    distributions = _compute_boarding_distribution(detailed_trips)
    if verbose:
        print("\nBoarding distribution (from detailed days):")
        for direction, stations in distributions.items():
            print(f"  {direction}:")
            for station, (b_frac, a_frac) in stations.items():
                print(f"    {station:<12} board={b_frac:.1%}  alight={a_frac:.1%}")

    # Step 3: Parse total-only days (30 dates, distribute using pattern)
    estimated_trips = _parse_total_only_sheet(wb, distributions)
    n_estimated_dates = len(set(k.rsplit("_", 1)[0] for k in estimated_trips))
    if verbose:
        print(f"\nTotal-only days: {n_estimated_dates} dates, "
              f"{len(estimated_trips)} trips (estimated distribution)")

    # Step 4: Connect to DuckDB and match
    db = duckdb.connect(DUCKDB_PATH)

    # Clear existing ridership data
    db.execute("DELETE FROM segment_ridership")
    if verbose:
        print("\nCleared existing segment_ridership data")

    # Get segments grouped by date + direction
    segments_by_key = _get_segments_by_date_direction(db)
    if verbose:
        print(f"Segments loaded: {sum(len(v) for v in segments_by_key.values())} "
              f"across {len(segments_by_key)} date-direction groups")

    # Match detailed days (observed data)
    print("\nMatching ridership to segments...")
    m1, u1 = _match_and_insert(db, detailed_trips, segments_by_key, "Observed (detailed)")

    # Match total-only days (estimated data)
    m2, u2 = _match_and_insert(db, estimated_trips, segments_by_key, "Estimated (totals)")

    total_matched = m1 + m2
    total_unmatched = u1 + u2

    # Verify
    count = db.execute("SELECT COUNT(*) FROM segment_ridership").fetchone()[0]
    non_null = db.execute(
        "SELECT COUNT(*) FROM segment_ridership WHERE passengers_on_board IS NOT NULL"
    ).fetchone()[0]

    # Stats on inserted data
    stats = db.execute("""
        SELECT
            MIN(passengers_on_board) as min_pax,
            MAX(passengers_on_board) as max_pax,
            AVG(passengers_on_board) as avg_pax,
            COUNT(*) as n
        FROM segment_ridership
    """).fetchone()

    db.close()

    if verbose:
        print(f"\nResults:")
        print(f"  Segments with ridership: {count} / 861 ({count/861*100:.1f}%)")
        print(f"  Matched: {total_matched}, Unmatched: {total_unmatched}")
        print(f"  Passengers — min: {stats[0]}, max: {stats[1]}, avg: {stats[2]:.1f}")
        print(f"\nPrevious: 0 segments had ridership (table was empty)")
        print(f"Now: {count} segments have real/estimated ridership data")

    return {
        "detailed_dates": n_detailed_dates,
        "estimated_dates": n_estimated_dates,
        "matched_observed": m1,
        "matched_estimated": m2,
        "unmatched": total_unmatched,
        "total_in_table": count,
        "pax_min": stats[0],
        "pax_max": stats[1],
        "pax_avg": round(stats[2], 1),
    }


if __name__ == "__main__":
    extract()
